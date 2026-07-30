"""UTF-8 CSV and accessible Excel exports."""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from web_scraper.models import COLUMNS, PersonRecord


def _frame(records: list[PersonRecord]) -> pd.DataFrame:
    return pd.DataFrame([record.as_dict() for record in records], columns=COLUMNS)


def export_rows(
    rows: Sequence[Mapping[str, str]],
    output: str | Path,
    columns: Sequence[str],
    sheet_name: str = "Scraped data",
) -> Path:
    """Export ordered rows to UTF-8 CSV or accessible Excel without merged cells."""
    path = Path(output)
    path.parent.mkdir(parents=True, exist_ok=True)
    frame = pd.DataFrame([dict(row) for row in rows], columns=list(columns)).fillna("")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        frame.to_csv(path, index=False, encoding="utf-8")
    elif suffix in {".xlsx", ".xlsm"}:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            frame.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.book[sheet_name]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            header_fill = PatternFill("solid", fgColor="1F4E78")
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
            for column_index, column in enumerate(frame.columns, start=1):
                values = [str(value) for value in frame[column].fillna("")]
                width = min(max([len(column), *(len(value) for value in values)]) + 2, 50)
                worksheet.column_dimensions[get_column_letter(column_index)].width = max(width, 12)
    else:
        raise ValueError("Output extension must be .csv, .xlsx, or .xlsm")
    return path


def export_records(records: list[PersonRecord], output: str | Path) -> Path:
    """Export the framework's default record shape."""
    return export_rows([record.as_dict() for record in records], output, COLUMNS)
